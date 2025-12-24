from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import List, Dict, Any

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

KEY_COLS = ["Year", "Quarter", "Region"]
VAL_COLS = ["Index", "Up To 5 Years Old Index", "Over 5 Years Old Index"]

@dataclass
class ExcelUpdateResult:
    rows_before: int
    rows_after: int
    updated_cells: int
    new_rows: int
    report_df: pd.DataFrame
    updated_df: pd.DataFrame

def _period_num(df: pd.DataFrame) -> pd.Series:
    return df["Year"].astype(int) * 10 + df["Quarter"].astype(int)

def compare_and_update_excel(
    db_excel_path: Path,
    sheet: str,
    extracted_df: pd.DataFrame,
    out_excel_path: Path,
    report_csv_path: Path,
    *,
    prevent_older_than_db: bool = True,
    float_tol: float = 1e-9
) -> ExcelUpdateResult:
    if not isinstance(db_excel_path, Path):
        db_excel_path = Path(db_excel_path)

    if not db_excel_path.exists():
        raise FileNotFoundError(f"DB Excel not found: {db_excel_path}")

    df_db = pd.read_excel(db_excel_path, sheet_name=sheet)
    if df_db.empty:
        raise ValueError("DB Excel sheet is empty.")

    df_db["Region"] = df_db["Region"].astype(str).str.strip()
    df_new = extracted_df.copy()
    df_new["Region"] = df_new["Region"].astype(str).str.strip()

    if prevent_older_than_db:
        min_period = int(_period_num(df_db).min())
        df_new = df_new[_period_num(df_new) >= min_period].copy()

    db_idx = df_db.set_index(KEY_COLS)
    new_idx = df_new.set_index(KEY_COLS)

    common = db_idx.index.intersection(new_idx.index)
    only_new = new_idx.index.difference(db_idx.index)

    changes: List[Dict[str, Any]] = []
    updated_cells = 0

    for k in common:
        for c in VAL_COLS:
            old = db_idx.loc[k, c]
            new = new_idx.loc[k, c]

            if pd.isna(old) and pd.isna(new):
                continue

            different = False
            try:
                different = (pd.isna(old) != pd.isna(new)) or abs(float(old) - float(new)) > float_tol
            except Exception:
                different = str(old) != str(new)

            if different:
                updated_cells += 1
                changes.append({
                    "ChangeType": "UPDATE",
                    "Year": k[0], "Quarter": k[1], "Region": k[2],
                    "Field": c,
                    "OldValue": old,
                    "NewValue": new
                })
                db_idx.loc[k, c] = new

    df_added = new_idx.loc[only_new].reset_index()
    for _, r in df_added.iterrows():
        changes.append({
            "ChangeType": "ADD_ROW",
            "Year": r["Year"], "Quarter": r["Quarter"], "Region": r["Region"],
            "Field": "",
            "OldValue": "",
            "NewValue": f'Index={r["Index"]}, UpTo5={r["Up To 5 Years Old Index"]}, Over5={r["Over 5 Years Old Index"]}'
        })

    df_updated = pd.concat([db_idx.reset_index(), df_added], ignore_index=True)
    df_updated = df_updated.sort_values(["Year", "Quarter", "Region"]).reset_index(drop=True)

    report_df = pd.DataFrame(changes)
    report_csv_path.parent.mkdir(parents=True, exist_ok=True)
    report_df.to_csv(report_csv_path, index=False)

    out_excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(db_excel_path)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet]
    ws.delete_rows(1, ws.max_row)

    for r in dataframe_to_rows(df_updated, index=False, header=True):
        ws.append(r)

    wb.save(out_excel_path)

    return ExcelUpdateResult(
        rows_before=len(df_db),
        rows_after=len(df_updated),
        updated_cells=updated_cells,
        new_rows=len(df_added),
        report_df=report_df,
        updated_df=df_updated,
    )
